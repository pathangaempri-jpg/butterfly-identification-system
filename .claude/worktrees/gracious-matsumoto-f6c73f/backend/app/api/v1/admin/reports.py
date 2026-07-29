"""Admin — PDF and Excel report exports — via Supabase PostgREST."""
import io
from datetime import datetime

from flask import Blueprint, send_file, request

from app.supabase_client import get_supabase
from app.utils.decorators import admin_required

admin_reports_bp = Blueprint("admin_reports", __name__)

_OBS_REPORT_SELECT = (
    "*, india_states(name), india_districts(name), species(common_name), "
    "users!observations_user_id_fkey(username, email)"
)


def _one(rel):
    if isinstance(rel, list):
        return rel[0] if rel else None
    return rel


def _get_date_filters():
    """Parse optional ?from_date=YYYY-MM-DD&to_date=YYYY-MM-DD from query string."""
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")
    result = {}
    if from_date:
        try:
            result["from"] = datetime.strptime(from_date, "%Y-%m-%d").isoformat()
        except ValueError:
            pass
    if to_date:
        try:
            result["to"] = datetime.strptime(to_date, "%Y-%m-%d").isoformat()
        except ValueError:
            pass
    return result


def _fetch_observations(limit: int) -> list:
    sb = get_supabase()
    dates = _get_date_filters()
    query = sb.table("observations").select(_OBS_REPORT_SELECT).eq("is_active", True)
    if dates.get("from"):
        query = query.gte("created_at", dates["from"])
    if dates.get("to"):
        query = query.lte("created_at", dates["to"])
    return query.order("created_at", desc=True).limit(limit).execute().data


def _fmt_date(value, fmt="%Y-%m-%d"):
    if not value:
        return ""
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).strftime(fmt)
    except (ValueError, AttributeError):
        return str(value)


# ── Observations PDF ───────────────────────────────────────────────────────────

@admin_reports_bp.get("/observations/pdf")
@admin_required
def observations_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet

    observations = _fetch_observations(1000)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Butterfly Observations Report", styles["Title"]))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    headers = ["#", "User", "State", "Species", "Status", "Date", "Privacy"]
    rows = [headers]
    for i, obs in enumerate(observations, 1):
        user = _one(obs.get("users"))
        state = _one(obs.get("india_states"))
        species = _one(obs.get("species"))
        rows.append([
            str(i),
            user["username"] if user else "N/A",
            state["name"] if state else str(obs.get("state_id")),
            species["common_name"] if species else "Unidentified",
            obs.get("verification_status"),
            _fmt_date(obs.get("observed_at")),
            obs.get("privacy"),
        ])

    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D6A4F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F0F0")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)

    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"observations_{datetime.utcnow().strftime('%Y%m%d')}.pdf",
    )


# ── Observations Excel ─────────────────────────────────────────────────────────

@admin_reports_bp.get("/observations/excel")
@admin_required
def observations_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    observations = _fetch_observations(5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Observations"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
    headers = [
        "ID", "Username", "Email", "State", "District", "Latitude", "Longitude",
        "Species", "Status", "Privacy", "Count", "Weather", "Activity", "Observed At",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for obs in observations:
        user = _one(obs.get("users"))
        state = _one(obs.get("india_states"))
        district = _one(obs.get("india_districts"))
        species = _one(obs.get("species"))
        ws.append([
            obs["id"],
            user["username"] if user else "",
            user["email"] if user else "",
            state["name"] if state else str(obs.get("state_id")),
            district["name"] if district else "",
            obs.get("latitude"),
            obs.get("longitude"),
            species["common_name"] if species else "",
            obs.get("verification_status"),
            obs.get("privacy"),
            obs.get("count_observed"),
            obs.get("weather") or "",
            obs.get("butterfly_activity") or "",
            _fmt_date(obs.get("observed_at"), "%Y-%m-%d %H:%M"),
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"observations_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
    )


# ── Users PDF ──────────────────────────────────────────────────────────────────

@admin_reports_bp.get("/users/pdf")
@admin_required
def users_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet

    sb = get_supabase()
    users = (
        sb.table("users").select("*, roles(name)").eq("is_active", True)
        .order("created_at", desc=True).limit(1000).execute().data
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("User Report", styles["Title"]),
        Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]),
        Spacer(1, 12),
    ]

    rows = [["#", "Username", "Full Name", "Role", "Verified", "Suspended", "Joined"]]
    for i, u in enumerate(users, 1):
        role = _one(u.get("roles"))
        rows.append([
            str(i), u["username"], u["full_name"],
            role["name"] if role else "N/A",
            "Yes" if u.get("is_verified") else "No",
            "Yes" if u.get("is_suspended") else "No",
            _fmt_date(u.get("created_at")),
        ])

    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#E3F2FD")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return send_file(
        buf, mimetype="application/pdf", as_attachment=True,
        download_name=f"users_{datetime.utcnow().strftime('%Y%m%d')}.pdf",
    )


# ── Species PDF ────────────────────────────────────────────────────────────────

@admin_reports_bp.get("/species/pdf")
@admin_required
def species_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet

    sb = get_supabase()
    all_species = (
        sb.table("species").select("*").eq("is_active", True)
        .order("family").order("common_name").limit(10000).execute().data
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Species Catalogue", styles["Title"]),
        Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]),
        Spacer(1, 12),
    ]

    rows = [["#", "Common Name", "Scientific Name", "Family", "Conservation", "Migratory", "Wingspan (mm)"]]
    for i, sp in enumerate(all_species, 1):
        rows.append([
            str(i),
            sp["common_name"], sp["scientific_name"], sp["family"],
            sp.get("conservation_status"),
            "Yes" if sp.get("is_migratory") else "No",
            f"{sp['wing_span_min_mm']}–{sp['wing_span_max_mm']}" if sp.get("wing_span_min_mm") else "",
        ])

    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4A148C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3E5F5")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return send_file(
        buf, mimetype="application/pdf", as_attachment=True,
        download_name=f"species_{datetime.utcnow().strftime('%Y%m%d')}.pdf",
    )
